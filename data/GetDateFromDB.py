import datetime, json, requests
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import ImageGrab
import win32com
import xlrd
from data.excel import *


class bolimlar():

    def get_path(self, *path):
        root_path = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(root_path, *path)

    def getimage(self, list, num):
        print('c')
        harf = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
        list1 = []
        a = len(list) // num
        s = -1
        c = 0
        for v in range(a):
            list1.append([])
            c += num
            s += 1
            for i in range(c - num, c):
                list1[s].append(list[i])

        import xlsxwriter

        with xlsxwriter.Workbook('data/test.xlsx') as workbook:
            worksheet = workbook.add_worksheet()

            for row_num, data in enumerate(list1):
                worksheet.write_row(row_num, 0, data)

        # ------------------------------------------------------------------------
        print('v')

        s = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        color = ['F7FC6D', '6ABDF6', 'FA5050', '74FE74', '4D8380', 'F1A15F']

        wb = load_workbook('data/test.xlsx')
        fill_pattern = PatternFill(patternType='solid', fgColor='6ABDF6')
        x = 2
        ws = wb['Sheet1']
        for i in s:
            length = 0
            for v in range(1, a + 1):
                length_of_column = len(str(ws[f'{i}{v}'].value))
                ws[f'{i}1'].fill = fill_pattern
                if length_of_column > length:
                    ws.column_dimensions[f'{i}'].width = length_of_column + x
                else:
                    ws.column_dimensions[f'{i}'].width = length + x
                for z in range(1, a + 1):
                    length_of_column1 = len(str(ws[f'{i}{z}'].value))
                    if length < length_of_column1:
                        length += length_of_column

        wb.save('sample1.xlsx')
        wb.save('data/text1.xlsx')

        # -----------------------------------------------------------------------

        # -----------------------------------------------------------------------
        print('b')

        pathfile = self.get_path('text1.xlsx')
        print('z')

        client = win32com.client.Dispatch("Excel.Application")
        # client = xlrd.open_workbook("sales.xlsx")
        print('a')

        wb = client.Workbooks.Open(pathfile)
        # wb = xlrd.open_workbook('text1.xlsx')

        ws = wb.Worksheets("Sheet1")

        ws.Range(f"A1:{harf[num - 1]}{a}").CopyPicture(Format=2)
        print('m')

        img = ImageGrab.grabclipboard()
        img.save(self.get_path('image.jpg'))
        wb.Close()
        client.Quit()
        a = 'image.jpg'
        print('n')

        return a

    def Darsjadvali(self, login, parol):
        # obj1=image.DJ()
        obj1 = excel()

        def lessondate(lesson_date):
            timestamp = datetime.datetime.fromtimestamp(lesson_date)
            return (timestamp.strftime('%Y-%m-%d'))

        lesson = ''
        count = 0

        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        group = (response.json()['data']['group']['id'])
        faculty = (response.json()['data']['faculty']['id'])
        semester = (response.json()['data']['semester']['code'])
        education_year = (response.json()['data']['semester']['education_year']['code'])
        page = 1
        url = f"https://student.urdu.uz/rest/v1/data/schedule-list?_faculty={faculty}&_group={group}&_semester={semester}&_education_year={education_year}&page={page}"

        payload = {}
        headers = {
            'Authorization': 'Bearer t9n6KJ9sgXhyI9A8cqXxOhuUlK6V5eHV',
            'Content-Type': 'application/json'
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        print('b')
        pageCount = response.json()['data']['pagination']['pageCount']
        today = datetime.date.today().strftime("%Y-%m-%d")
        lessonlist2 = []
        lessonlist2.append('Sana')
        lessonlist2.append('')
        lessonlist2.append('Fan nomi')
        lessonlist2.append('Fan turi')
        lessonlist2.append("O'qituchi")
        lessonlist2.append('Dars vaqti')
        lessonlist2.append('Xona')
        for i in range(pageCount):
            url = f"https://student.urdu.uz/rest/v1/data/schedule-list?_faculty={faculty}&_group={group}&_semester={semester}&_education_year={education_year}&page={i}"
            response = requests.request("GET", url, headers=headers, data=payload)
            # response 200 qaytganini tekshirish kerak;
            items = response.json()['data']['items']
            for item in items:
                _week = item['_week']
                _weekStart = lessondate(item['weekStartTime'])
                _weekEnd = lessondate(item['weekEndTime'])
                if _weekStart <= today <= _weekEnd:
                    if today == lessondate(item['lesson_date']):
                        count += 1
                        if f"{lessondate(item['lesson_date'])}" not in lessonlist2:
                            lessonlist2.append(f"{lessondate(item['lesson_date'])}")
                            lessonlist2.append(f"{count}")
                            lessonlist2.append(f"{item['subject']['name']}")
                            lessonlist2.append(f"{item['trainingType']['name']}")
                            lessonlist2.append(f"{item['employee']['name']}")
                            lessonlist2.append(f"{item['lessonPair']['start_time']}-{item['lessonPair']['end_time']}")
                            lessonlist2.append(f"{item['auditorium']['name']}")

                        else:
                            lessonlist2.append(f"")
                            lessonlist2.append(f"{count}")
                            lessonlist2.append(f"{item['subject']['name']}")
                            lessonlist2.append(f"{item['trainingType']['name']}")
                            lessonlist2.append(f"{item['employee']['name']}")
                            lessonlist2.append(f"{item['lessonPair']['start_time']}-{item['lessonPair']['end_time']}")
                            lessonlist2.append(f"{item['auditorium']['name']}")

        if len(lessonlist2) == 7:
            for n in range(6):
                lessonlist2.remove(lessonlist2[0])
            n = ['Bugun', 'sizga', 'dars', "yo'q", '', '']
            for i in n:
                lessonlist2.append(i)

        img = self.getimage(lessonlist2, 7)
        print(img)

        return img

    def Darsjadvali1(self, login, parol):

        def lessondate(lesson_date):
            timestamp = datetime.datetime.fromtimestamp(lesson_date)
            return (timestamp.strftime('%Y-%m-%d'))

        lesson1 = ''
        count = 0

        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        group = (response.json()['data']['group']['id'])
        faculty = (response.json()['data']['faculty']['id'])
        semester = (response.json()['data']['semester']['code'])
        education_year = (response.json()['data']['semester']['education_year']['code'])
        page = 1
        url = f"https://student.urdu.uz/rest/v1/data/schedule-list?_faculty={faculty}&_group={group}&_semester={semester}&_education_year={education_year}&page={page}"

        payload = {}
        headers = {
            'Authorization': 'Bearer t9n6KJ9sgXhyI9A8cqXxOhuUlK6V5eHV',
            'Content-Type': 'application/json'
        }
        response = requests.request("GET", url, headers=headers, data=payload)
        pageCount = response.json()['data']['pagination']['pageCount']
        today = datetime.date.today().strftime("%Y-%m-%d")
        currentWeek = 0
        lessonlist = []
        lessonlist.append('Sana')
        lessonlist.append('')
        lessonlist.append('Fan nomi')
        lessonlist.append('Fan turi')
        lessonlist.append("O'qituchi")
        lessonlist.append('Dars vaqti')
        lessonlist.append('Xona')
        for i in range(pageCount):
            url = f"https://student.urdu.uz/rest/v1/data/schedule-list?_faculty={faculty}&_group={group}&_semester={semester}&_education_year={education_year}&page={i}"
            response = requests.request("GET", url, headers=headers, data=payload)
            # response 200 qaytganini tekshirish kerak;
            items = response.json()['data']['items']
            for item in items:
                _week = item['_week']
                _weekStart = lessondate(item['weekStartTime'])
                _weekEnd = lessondate(item['weekEndTime'])
                if _weekStart <= today <= _weekEnd:
                    currentWeek = _week
                if currentWeek == _week:
                    count += 1
                    if f"{lessondate(item['lesson_date'])}" not in lessonlist:
                        count *= 0
                        count += 1
                        lessonlist.append(f"{lessondate(item['lesson_date'])}")
                        lessonlist.append(f"{count}")
                        lessonlist.append(f"{item['subject']['name']}")
                        lessonlist.append(f"{item['trainingType']['name']}")
                        lessonlist.append(f"{item['employee']['name']}")
                        lessonlist.append(f"{item['lessonPair']['start_time']}-{item['lessonPair']['end_time']}")
                        lessonlist.append(f"{item['auditorium']['name']}")
                    else:
                        lessonlist.append('')
                        lessonlist.append(f'{count}')
                        lessonlist.append(f"{item['subject']['name']}")
                        lessonlist.append(f"{item['trainingType']['name']}")
                        lessonlist.append(f"{item['employee']['name']}")
                        lessonlist.append(f"{item['lessonPair']['start_time']}-{item['lessonPair']['end_time']}")
                        lessonlist.append(f"{item['auditorium']['name']}")

        if len(lessonlist) == 7:
            for n in range(6):
                lessonlist.remove(lessonlist[0])
            n = ['Bu', 'haftada', 'dars', "belgilanmagan", 'yoki', 'yakshanba']
            for i in n:
                lessonlist.append(i)
        img = self.getimage(lessonlist, 7)

        return img

    def Davomat(self, login, parol):

        def lessondate(lesson_date):
            timestamp = datetime.datetime.fromtimestamp(lesson_date)
            return (timestamp.strftime('%Y-%m-%d'))

        davomat = []
        davomat.append('')
        davomat.append('Sana')
        davomat.append('Vaqti')
        davomat.append('Fan nomi')
        davomat.append('Fan turi')
        davomat.append("O'qituvchi")
        a = 1
        s = ''
        count = 0

        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        semester = response.json()['data']['semester']['code']

        url = f"https://student.urdu.uz/rest/v1/education/attendance?_semester={semester}"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'text/plain'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        davomatlist = response.json()['data']

        for i in davomatlist:
            if semester == i['semester']['code']:
                davomat.append(f'{a}')
                a += 1
                davomat.append(lessondate(i['lesson_date']))
                davomat.append(i['lessonPair']['start_time'])
                davomat.append(i['subject']['name'])
                davomat.append(i['trainingType']['name'])
                davomat.append(i['employee']['name'])

        if len(davomat) == 6:
            for n in range(6):
                davomat.remove(davomat[0])
            n = ["ma'lumot", 'mavjud', 'emas', "", '', '']
            for i in n:
                davomat.append(i)

        img = self.getimage(davomat, 6)

        return img

    def Buyruqlar(self, login, parol):
        # obj1=image()
        s = ''
        count = 0
        a = 1
        hujjat = []
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/student/document-all"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        hujjat.append('1')
        hujjat.append(response.json()['data'][::-1][0]['name'])
        hujjat.append(response.json()['data'][::-1][0]['attributes'][2]['value'])
        hujjat.append('Kursdan kursga ko‘chirish to‘g‘risida')
        hujjat.append(response.json()['data'][::-1][0]['attributes'][1]['value'])

        hujjat.append(response.json()['data'][::-1][1]['name'])
        hujjat.append(response.json()['data'][::-1][1]['attributes'][2]['value'])
        hujjat.append('Kursdan kursga ko‘chirish to‘g‘risida')
        hujjat.append(response.json()['data'][::-1][1]['attributes'][1]['value'])

        if len(hujjat) == 0:
            s += "Ma'lumot mavjud emas"
        else:
            for i in hujjat:
                count += 1
                s += str(i)
                s += '      '
                if count % 5 == 0:
                    s += '\n'
                    a += 1
                    s += str(a)
                    s += '       '
        return s

    def Malumotlar(self, login, parol):
        malumot = []
        s = ''
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        token = response.json()['data']['token']
        url = "https://student.urdu.uz/rest/v1/student/reference"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        malumot.append(response.json()['data'][0]['reference_number'])
        malumot.append(response.json()['data'][0]['semester']['education_year']['name'])
        malumot.append(response.json()['data'][0]['level']['name'])
        malumot.append(response.json()['data'][0]['semester']['name'])

        for i in malumot:
            s += str(i)
            s += '    '
        return s

    def NazoratjadvaliY(self, login, parol):
        def lessondate(lesson_date):
            timestamp = datetime.datetime.fromtimestamp(lesson_date)
            return (timestamp.strftime('%Y-%m-%d'))

        s1 = ''
        count = 0

        nazoratjadvali1 = []
        a = 1
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        semester = response.json()['data']['semester']['code']

        url = f"https://student.urdu.uz/rest/v1/education/exam-table?semester={semester}"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)

        nazorat = response.json()['data']

        for i in nazorat:
            if semester == i['semester']['code'] and i['examType']['name'] != 'Oraliq nazorat':
                nazoratjadvali1.append(a)
                a += 1
                nazoratjadvali1.append(i['examType']['name'])
                nazoratjadvali1.append(i['subject']['name'])
                nazoratjadvali1.append(i['semester']['name'])
                nazoratjadvali1.append(i['employee']['name'])
                nazoratjadvali1.append(i['auditorium']['name'])
                nazoratjadvali1.append(i['lessonPair']['start_time'])
                nazoratjadvali1.append(lessondate(i['examDate']))
                nazoratjadvali1.append(i['educationYear']['current'])

        if len(nazoratjadvali1) == 0:
            s1 += "Ma'lumot mavjud emas"
        else:
            for i in nazoratjadvali1:
                count += 1
                s1 += str(i)
                s1 += '  '
                if count % 9 == 0:
                    s1 += '\n'
        return s1

    def NazoratjadvaliO(self, login, parol):
        def lessondate(lesson_date):
            timestamp = datetime.datetime.fromtimestamp(lesson_date)
            return (timestamp.strftime('%Y-%m-%d'))

        s = ''
        count = 0

        nazoratjadvali = []
        a = 1
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        semester = response.json()['data']['semester']['code']

        url = f"https://student.urdu.uz/rest/v1/education/exam-table?semester={semester}"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)

        nazorat = response.json()['data']

        for i in nazorat:
            if semester == i['semester']['code'] and i['examType']['name'] == 'Oraliq nazorat':
                nazoratjadvali.append(a)
                a += 1
                nazoratjadvali.append(i['examType']['name'])
                nazoratjadvali.append(i['subject']['name'])
                nazoratjadvali.append(i['semester']['name'])
                nazoratjadvali.append(i['employee']['name'])
                nazoratjadvali.append(i['auditorium']['name'])
                nazoratjadvali.append(i['lessonPair']['start_time'])
                nazoratjadvali.append(lessondate(i['examDate']))
                nazoratjadvali.append(i['educationYear']['current'])

        if len(nazoratjadvali) == 0:
            s += "Ma'lumot mavjud emas"
        else:
            for i in nazoratjadvali:
                count += 1
                s += str(i)
                s += '  '
                if count % 9 == 0:
                    s += '\n'
        return s

    def Oquvreja(self, login, parol):
        javob = []
        fanlar = []
        fanlar.append('')
        fanlar.append('Fanlar')
        fanlar.append('Fan turi')
        fanlar.append('Yuklama')
        fanlar.append('Kredit')
        s = ''
        count = 0

        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}'
        }

        response1 = requests.request("GET", url, headers=headers, data=payload)
        # javob.append(response.json()['data']['semester']['name'])
        # print(javob)

        url = "https://student.urdu.uz/rest/v1/education/performance"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        a = response.json()['data'][::-1][0]['_semester']
        b = response.json()['data']
        for i in b:
            if a == i['_semester']:
                count += 1
                fanlar.append(count)
                fanlar.append(i['subject']['name'])
                fanlar.append(i['subjectType']['name'])
                fanlar.append(i['total_acload'])
                fanlar.append(float(i['credit']))

        if len(fanlar) == 5:
            for n in range(6):
                fanlar.remove(fanlar[0])
            n = ["Ma'lumot", 'mavjud', 'emas', "", '', '']
            for i in n:
                fanlar.append(i)

        img = self.getimage(fanlar, 5)

        return img

    def Ozlashtirish(self, login, parol):

        s = ''
        count = 1
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        token = response.json()['data']['token']
        url = "https://student.urdu.uz/rest/v1/education/performance"
        fanlar1 = []
        fanlar1.append('')
        fanlar1.append('Fanlar')
        fanlar1.append('On')
        fanlar1.append("Yn")
        fanlar1.append("Umumiy")

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        a = response.json()['data'][::-1][0]['_semester']
        b = response.json()['data']

        for i in b:
            if a == i['_semester']:
                fanlar1.append(count)
                count += 1
                fanlar1.append(i['subject']['name'])
                for ii in i['performances']:
                    fanlar1.append(ii['grade'])

        if len(fanlar1) == 5:
            for n in range(4):
                fanlar1.remove(fanlar1[0])
            n = ["Ma'lumot", 'mavjud', 'emas', '', "", '']
            for i in n:
                fanlar1.append(i)

        img = self.getimage(fanlar1, 5)

        return img

    def Reytingdaftarcha(self, login, parol):
        reyting = []
        reyting.append('')
        reyting.append('Fanlar')
        reyting.append('Fan turi')
        reyting.append('Yuklama')
        reyting.append('Kredit')
        s = ''
        count = 0
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        token = response.json()['data']['token']
        url = "https://student.urdu.uz/rest/v1/education/performance"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        print(response)
        a = response.json()['data'][::-1][0]['_semester']
        b = response.json()['data']
        for i in b:
            if a == i['_semester']:
                count += 1
                reyting.append(count)
                reyting.append(i['subject']['name'])
                reyting.append(i['subjectType']['name'])
                reyting.append(i['total_acload'])
                reyting.append(float(i['credit']))

        if len(reyting) == 5:
            for n in range(4):
                reyting.remove(reyting[0])
            n = ["Ma'lumot", 'mavjud', 'emas', '', "", '']
            for i in n:
                reyting.append(i)

        img = self.getimage(reyting, 5)

        return img

    def Talabahujjati(self, login, parol):
        javobb = []
        count = 0
        s = ''
        url = "https://student.urdu.uz/rest/v1/auth/login"

        payload = json.dumps({
            "login": f"{login}",
            "password": f"{parol}"
        })
        headers = {
            'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJ2MVwvYXV0aFwvbG9naW4iLCJhdWQiOiJ2MVwvYXV0aFwvbG9naW4iLCJleHAiOjE2NjkzNTY5ODIsImp0aSI6IjM0MTIwMTEwOTYwNSIsInN1YiI6IjM1NTg4In0.d91veAwleMVeehIzzSG_KxG1vNPLVYe7aeq_ozoWHJc',
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)
        token = response.json()['data']['token']

        url = "https://student.urdu.uz/rest/v1/account/me"

        payload = {}
        headers = {
            'Authorization': f'Bearer {token}'
        }

        response = requests.request("GET", url, headers=headers, data=payload)
        javobb.append("O'QUV VARAQA")
        javobb.append('')
        javobb.append('Kurs')
        javobb.append(response.json()['data']['level']['name'])
        javobb.append("O'quv yili")
        javobb.append(response.json()['data']['semester']['education_year']['name'])
        javobb.append('Semester')
        javobb.append(response.json()['data']['semester']['name'])
        javobb.append('REYTING DAFTARCHA')
        javobb.append('')
        javobb.append('Mutaxassislik')
        javobb.append(response.json()['data']['specialty']['name'])
        javobb.append('Semester')
        javobb.append(response.json()['data']['semester']['name'])
        javobb.append('Kurs')
        javobb.append(response.json()['data']['level']['name'])

        if len(javobb) == 0:
            for n in range(4):
                javobb.remove(javobb[0])
            n = ["Ma'lumot", 'mavjud', 'emas', '', "", '']
            for i in n:
                javobb.append(i)

        img = self.getimage(javobb, 2)
        return img
